import subprocess
import openpyxl

# Define as configurações que serão testadas
configurations = [
    ("8", "0"), ("9", "0"), ("10", "0"),
    ("8", "2"), ("9", "2"), ("10", "2"),
    ("8", "4"), ("9", "4"), ("10", "4"),
    ("8", "8"), ("9", "8"), ("10", "8")
]

# Inicializa uma nova planilha Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# Define os cabeçalhos da planilha
sheet.cell(row=1, column=1, value="N")
sheet.cell(row=1, column=2, value="Threads")
sheet.cell(row=1, column=3, value="Tempo")

# Itera sobre as configurações e executa o código para cada uma delas
for i, config in enumerate(configurations):
    n, threads = config
    print(f"Testando configuração {i+1} de {len(configurations)}: {n} rainhas, {threads} threads")
    
    # Executa o código e captura o tempo de execução
    output = subprocess.run(["mpirun", "-np", "1", "nrainhashibrido", n, threads], capture_output=True, text=True)
    time = output.stdout.strip()
    
    # Adiciona os resultados na planilha
    row = i + 2
    sheet.cell(row=row, column=1, value=n)
    sheet.cell(row=row, column=2, value=threads)
    sheet.cell(row=row, column=3, value=time)

# Salva a planilha em um arquivo Excel
workbook.save("resultados.xlsx")
